from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render,redirect
from django.contrib.auth.models import User, auth
from django.utils.text import capfirst
from django.contrib import messages
from . models import *
import json
from django.http.response import JsonResponse
from django.utils.crypto import get_random_string
from datetime import date
from datetime import timedelta
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.template.response import TemplateResponse


# Create your views here.
def home(request):
  return render(request, 'home.html')



    
def homepage(request):
  com =  company.objects.get(user = request.user)
  allmodules= modules_list.objects.get(company=com.id,status='New')
  context = {
              'company' : com,
              'allmodules':allmodules
          }
  return render(request, 'company/homepage.html', context)

# @login_required(login_url='login')
def staffhome(request):
  if 'staff_id' in request.session:
    if request.session.has_key('staff_id'):
      staff_id = request.session['staff_id']
           
    else:
      return redirect('/')
  staff =  staff_details.objects.get(id=staff_id)

  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request, 'staff/staffhome.html', context)




# @login_required(login_url='login')
def logout(request):
    auth.logout(request)
    return redirect('/')

def view_profile(request):
  com =  company.objects.get(user = request.user) 
  selected_options = request.session.get('selected_options', None)
  
  context = {
              'company' : com,
              'selected_options': json.dumps(selected_options)
          }
  return render(request,'profile.html',context)
  
def edit_profile(request,pk):
  com= company.objects.get(id = pk)
  user1 = User.objects.get(id = com.user_id)
  selected_options = request.session.get('selected_options', None)

  if request.method == "POST":

      user1.first_name = capfirst(request.POST.get('f_name'))
      user1.last_name  = capfirst(request.POST.get('l_name'))
      user1.email = request.POST.get('email')
      com.contact_number = request.POST.get('cnum')
      com.address = capfirst(request.POST.get('ards'))
      com.company_name = request.POST.get('comp_name')
      com.company_email = request.POST.get('comp_email')
      com.city = request.POST.get('city')
      com.state = request.POST.get('state')
      com.country = request.POST.get('country')
      com.pincode = request.POST.get('pinc')
      com.gst_num = request.POST.get('gst')
      com.pan_num = request.POST.get('pan')
      com.business_name = request.POST.get('bname')
      com.company_type = request.POST.get('comp_type')
      if len(request.FILES)!=0 :
          com.profile_pic = request.FILES.get('file')

      com.save()
      user1.save()
      return redirect('view_profile')

  context = {
      'company' : com,
      'user1' : user1,
      'selected_options': json.dumps(selected_options)
  } 
  return render(request,'company/edit_profile.html',context)


def sale_invoices(request):
  return render(request, 'sale_invoices.html')

def estimate_quotation(request):
  return render(request, 'estimate_quotation.html')

def payment_in(request):
  return render(request, 'payment_in.html')
    
def sale_order(request):
  return render(request, 'sale_order.html')

def delivery_chellan(request):
  return render(request, 'delivery_chellan.html')

def sale_return_cr(request):
  return render(request, 'sale_return_cr.html')


# created by athul
def settings(request):
  com =  company.objects.get(user = request.user)
  selected_options = request.session.get('selected_options', None)
  
  context = {
              'company' : com,
              'selected_options': json.dumps(selected_options),
              
          }
  return render(request, 'company/settings.html',context)

def hide_options(request):
    
    com =  company.objects.get(user = request.user)
    if request.method == 'POST':
        selected_options = list(request.POST.getlist('selected_options'))

    request.session['selected_options'] = selected_options
    
    context = {'selected_options': json.dumps(selected_options),
               'company' : com}
   
    return render(request, 'company/homepage.html', context)

# ------created by athul------

def company_reg(request):
  return render(request,'company/register.html')

def register(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    action = request.POST['r']
    did = request.POST['did']
    if did != '':
      if Distributors_details.objects.filter(distributor_id=did).exists():
        distributor = Distributors_details.objects.get(distributor_id=did)
      else :
          messages.info(request, 'Sorry, distributor id does not exists')
          return redirect('company_reg')
    

    
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('company_reg')
      

      elif not User.objects.filter(email = email_id).exists():
        
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        if did != '':
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action,Distributors=distributor)
          cust_data.save()
        
          return redirect('company_reg2')
        else:
          data = User.objects.get(id = user_data.id)
          cust_data = company( contact=mobile,
                             user = data,reg_action=action)
          cust_data.save()
        
          return redirect('company_reg2')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('company_reg')
    return render(request,'company/register.html')
    
    
def company_reg2(request):
  terms=payment_terms.objects.all()
  
  return render(request,'company/register2.html',{'terms':terms})  
def add_company(request):
  
  print(id)
  if request.method == 'POST':
    email=request.POST['email']
    user=User.objects.get(email=email)
    
    c =company.objects.get(user = user)
    c.company_name=request.POST['cname']

    c.address=request.POST['address']
    c.city=request.POST['city']
    c.state=request.POST['state']
    c.country=request.POST['country']
    c.pincode=request.POST['pincode']
    c.pan_number=request.POST['pannumber']
    c.gst_type=request.POST['gsttype']
    c.gst_no=request.POST['gstno']
    c.profile_pic=request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    c.dateperiod=terms
    c.start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    c.End_date=end


    
    

    code=get_random_string(length=6)
    if company.objects.filter(Company_code = code).exists():
       code2=get_random_string(length=6)
       c.Company_code=code2
    else:
      c.Company_code=code

   
    c.save()
    # messages.success(request, 'Welcome'+ ' ' +  user.first_name +' '+user.last_name + ' ')

    return redirect('Allmodule',user.id)  
  return render(request,'company/register2.html')   

def staff_register(request):
  com=company.objects.all()

  return render(request, 'staff/staffreg.html',{'company':com})

def staff_registraction(request):
  if request.method == 'POST':
    fn=request.POST['fname']
    ln=request.POST['lname']
    email=request.POST['eid']
    un=request.POST['uname']
    pas=request.POST['pass']
    ph=request.POST['ph']
    code=request.POST['code']
    com=company.objects.get(Company_code=code)
    img=request.FILES.get('image')

    if staff_details.objects.filter(user_name=un).exists():
      messages.info(request, 'Sorry, Username already exists')
      print("1")
      return redirect('staff_register')
    elif staff_details.objects.filter(email=email).exists():
      messages.info(request, 'Sorry, Email already exists')
      print("2")
      return redirect('staff_register')
    else:
      
      staff=staff_details(first_name=fn,last_name=ln,email=email,user_name=un,password=pas,contact=ph,img=img,company=com)
      staff.save()
      print("success")
      return redirect('log_page')

  else:
    print(" error")
    return redirect('staff_register')
  

def adminaccept(request,id):
  data=company.objects.filter(id=id).update(superadmin_approval=1)
  return redirect('client_request')
def adminreject(request,id):
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('client_request')



def companyaccept(request,id):
  data=staff_details.objects.filter(id=id).update(Action=1)
  return redirect('staff_request')

def companyreject(request,id):
  data=staff_details.objects.get(id=id)
  
  data.delete()
  return redirect('staff_request')

def client_request(request):
  data = company.objects.filter(superadmin_approval = 0,reg_action='self').order_by('-id')
  
  all = company.objects.filter(superadmin_approval = 1)
  return render(request,'admin/client_request.html',{'data': data,'all':all})

def client_request_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_request_overview.html',{'company':com,'allmodules':allmodules})

def client_details(request):
  data = company.objects.filter(superadmin_approval = 1,reg_action='self').order_by('-id')
  return render(request,'admin/client_details.html',{"data":data})
def client_details_overview(request,id): 
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  return render(request,'admin/client_details_overview.html',{'company':com,'allmodules':allmodules})

def staff_request(request):
  com =  company.objects.get(user = request.user)
  staff = staff_details.objects.filter(company=com,Action=0).order_by('-id')
  allmodules= modules_list.objects.get(company=com.id,status='New')
  return render(request,'company/staff_request.html',{'staff':staff,'company':com,'allmodules':allmodules}) 
 
def View_staff(request):
  com =  company.objects.get(user = request.user)
  staff = staff_details.objects.filter(company=com,Action=0)
  allstaff = staff_details.objects.filter(company=com,Action=1).order_by('-id')
  allmodules= modules_list.objects.get(company=com.id,status='New')

  return render(request, 'company/view_staff.html',{'staff':staff,'company':com,'allstaff':allstaff,'allmodules':allmodules})

def payment_term(request):
  terms = payment_terms.objects.all()
  
  return render(request,'admin/payment_terms.html',{'terms':terms})
def add_payment_terms(request):
  if request.method == 'POST':
    num=int(request.POST['num'])
    select=request.POST['select']
    if select == 'Years':
      days=int(num)*365
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')

    else:  
      days=int(num*30)
      pt = payment_terms(payment_terms_number = num,payment_terms_value = select,days = days)
      pt.save()
      messages.info(request, 'Payment term is added')
      return redirect('payment_term')


  return redirect('payment_term')

def Allmodule(request,uid):
  user=User.objects.get(id=uid)
  return render(request,'company/modules.html',{'user':user})

def addmodules(request,uid):
  if request.method == 'POST':
    com=company.objects.get(user=uid)
    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14)
    data.save()

    return redirect('log_page')
  


def companyreport(request):
  com =  company.objects.get(user = request.user)
  allmodules= modules_list.objects.get(company=com.id,status='New')
  return render(request,'company/companyreport.html',{'company':com,'allmodules':allmodules})  

def Companyprofile(request):
  com =  company.objects.get(user = request.user)
  allmodules= modules_list.objects.get(company=com.id,status='New')
  return render(request,'company/companyprofile.html',{'company':com,'allmodules':allmodules})

def editcompanyprofile(request):
  com =  company.objects.get(user = request.user)
  allmodules= modules_list.objects.get(company=com.id,status='New')
  terms=payment_terms.objects.all()
  return render(request,'company/editcompanyprofile.html',{'company':com,'allmodules':allmodules,'terms':terms})

def editcompanyprofile_action(request):
  com =  company.objects.get(user = request.user)
  if request.method == 'POST':
    com.company_name = request.POST['cname']
    com.user.email = request.POST['email']
    com.contact = request.POST['ph']
    com.address = request.POST['address']
    com.city = request.POST['city']
    com.state = request.POST['state']
    com.country = request.POST['country']
    com.pincode = request.POST['pincode']

    t = request.POST['select']
    terms = payment_terms.objects.get(id=t)
    com.dateperiod = terms
    com.start_date=date.today()
    days=int(terms.days)

    end= date.today() + timedelta(days=days)
    com.End_date=end

    old=com.profile_pic
    new=request.FILES.get('image')
    if old!=None and new==None:
      com.profile_pic=old
    else:
      com.profile_pic=new
    
    com.save() 
    com.user.save() 
    return redirect('Companyprofile') 



  return redirect('Companyprofile')

def editmodule(request):
  com =  company.objects.get(user = request.user)
  allmodules= modules_list.objects.get(company=com.id,status='New')
  return render(request,'company/editmodule.html',{'company':com,'allmodules':allmodules})

def editmodule_action(request):
  if request.method == 'POST':
    com = company.objects.get(user = request.user)
    # if modules_list.objects.filter(company=com.id,status='Old').exists():
    #   old=modules_list.objects.filter(company=com.id,status='Old')
    #   old.delete()

    # old_data=modules_list.objects.get(company=com.id,status='New')  
    # old_data.status='Old'
    # old_data.save()



    c1=request.POST.get('c1')
    c2=request.POST.get('c2')
    c3=request.POST.get('c3')
    c4=request.POST.get('c4')
    c5=request.POST.get('c5')
    c6=request.POST.get('c6')
    c7=request.POST.get('c7')
    c8=request.POST.get('c8')
    c9=request.POST.get('c9')
    c10=request.POST.get('c10')
    c11=request.POST.get('c11')
    c12=request.POST.get('c12')
    c13=request.POST.get('c13')
    c14=request.POST.get('c14')
    
    data=modules_list(company=com,sales_invoice = c1,
                      Estimate=c2,Payment_in=c3,sales_order=c4,
                      Delivery_challan=c5,sales_return=c6,Purchase_bills=c7,
                      Payment_out=c8,Purchase_order=c9,Purchase_return=c10,
                      Bank_account=c11,Cash_in_hand=c12, cheques=c13,Loan_account=c14,status='Pending')
    data.save()
    data1=modules_list.objects.filter(company=com.id,status='Pending').update(update_action=1)
    return redirect('Companyprofile')
    
    
  return redirect('Companyprofile')

def admin_notification(request):
  data= modules_list.objects.filter(update_action=1,status='Pending')

  return render(request,'admin/admin_notification.html',{'data':data})

def module_updation_details(request,mid):
  data= modules_list.objects.get(id=mid)
  allmodules= modules_list.objects.get(company=data.company,status='Pending')
  old_modules = modules_list.objects.get(company=data.company,status='New')

  return render(request,'admin/module_updation_details.html',{'data':data,'allmodules':allmodules,'old_modules':old_modules})

def module_updation_ok(request,mid):
  
  old=modules_list.objects.get(company=mid,status='New')
  old.delete()

  data=modules_list.objects.get(company=mid,status='Pending')  
  data.status='New'
  data.save()
  data1=modules_list.objects.filter(company=mid).update(update_action=0)
  return redirect('admin_notification')

def staff_profile(request,sid):
  staff =  staff_details.objects.get(id=sid)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/staff_profile.html',context)

def editstaff_profile(request,sid):
  staff =  staff_details.objects.get(id=sid)
  allmodules= modules_list.objects.get(company=staff.company,status='New')
  context = {
              'staff' : staff,
              'allmodules':allmodules

          }
  return render(request,'staff/editstaff_profile.html',context)

def editstaff_profile_action(request,sid):
  if request.method == 'POST':
    staff =  staff_details.objects.get(id=sid)
    staff.first_name = request.POST['fname']
    staff.last_name = request.POST['lname']
    staff.user_name = request.POST['uname']
    staff.email = request.POST['email']
    staff.contact = request.POST['ph']
    old=staff.img
    new=request.FILES.get('image')
    if old!=None and new==None:
      staff.img=old
    else:
      staff.img=new

    staff.save()  

    return redirect ('staff_profile',staff.id)
  return redirect ('staff_profile',staff.id)




def distributor_reg(request):
  terms=payment_terms.objects.all()
  return render(request,'distributor/distributor_reg.html',{'terms':terms})
def distributor_reg_action(request):
  if request.method == 'POST':
    first_name = request.POST['fname']
    last_name = request.POST['lname']
    user_name = request.POST['uname']
    email_id = request.POST['eid']
    mobile = request.POST['ph']
    passw = request.POST['pass']
    c_passw = request.POST['cpass']
    pic = request.FILES.get('image')

    select=request.POST['select']
    terms=payment_terms.objects.get(id=select)
    # c.dateperiod=terms
    start_date=date.today()
    days=int(terms.days)

    
    end= date.today() + timedelta(days=days)
    End_date=end

    code=get_random_string(length=6)
    if Distributors_details.objects.filter(distributor_id = code).exists():
       code=get_random_string(length=6)
  
    if passw == c_passw:
      if User.objects.filter(username = user_name).exists():
        messages.info(request, 'Sorry, Username already exists')
        return redirect('distributor_reg')
      

      elif not User.objects.filter(email = email_id).exists():
    
        user_data = User.objects.create_user(first_name = first_name,
                        last_name = last_name,
                        username = user_name,
                        email = email_id,
                        password = passw)
        user_data.save()
        
        data = User.objects.get(id = user_data.id)
        distributor_data = Distributors_details(contact=mobile,distributor_id=code,img=pic,
                                                payment_term=terms,start_date=start_date,End_date=End_date,
                                                user = data)
        distributor_data.save()
        
        return redirect('log_page')
      else:
        messages.info(request, 'Sorry, Email already exists')
        return redirect('distributor_reg')
  return render(request,'distributor/distributor_reg.html')
 
def distributor_home(request):
  distributor =  Distributors_details.objects.get(user = request.user)

  return render(request,'distributor/distributor_home.html',{'distributor':distributor})

# ----athul ----10-11-2023---

def log_page(request):
  return render(request, 'log.html')
  
def login(request):
  if request.method == 'POST':
    user_name = request.POST['username']
    passw = request.POST['password']
    
    log_user = auth.authenticate(username = user_name,
                                  password = passw)
    
    if log_user is not None:
      auth.login(request, log_user)
      if request.user.is_staff==1:
        return redirect('adminhome')
      else:
        if company.objects.filter(user=request.user).exists():
          data=company.objects.get(user=request.user)
          if data.superadmin_approval == 1 or data.Distributor_approval == 1:
             return redirect('homepage')
          else:
            messages.info(request, 'Approval is Pending..')
            return redirect('log_page')
        if Distributors_details.objects.filter(user=request.user).exists():
          data=Distributors_details.objects.get(user=request.user)
          if data.Log_Action == 1:
            return redirect('distributor_home')

          else:
            messages.info(request, 'Approval is Pending..')
            return redirect('log_page')


    elif staff_details.objects.filter(user_name=user_name,password=passw).exists(): 
      data=staff_details.objects.get(user_name=user_name,password=passw)
      if data.Action == 1:
        request.session["staff_id"]=data.id
        if 'staff_id' in request.session:
          if request.session.has_key('emp_id'):
            staff_id = request.session['staff_id']
            print(staff_id)
 
          return redirect('staffhome')  
      else:
        messages.info(request, 'Approval is Pending..')
        return redirect('log_page')
    else:
      messages.info(request, 'Invalid Username or Password. Try Again.')
      return redirect('log_page')
  else:  
   return redirect('log_page')
      
def clients(request):
  return render(request,'admin/clients.html')

def distributors(request):
  return render(request,'admin/distributors.html')  

def distributor_request(request):
  data = Distributors_details.objects.filter(Log_Action = 0).order_by('-id')
  return render(request,'admin/distributor_request.html',{'data':data})

def admin_distributor_accept(request,id):
  data=Distributors_details.objects.filter(id=id).update(Log_Action=1)
  return redirect('distributor_request')
def admin_distributor_reject(request,id):
  data=Distributors_details.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('distributor_request')

def distributor_request_overview(request,id):
  data=Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_request_overview.html',{'data':data})

def distributor_details(request):
  data = Distributors_details.objects.filter(Log_Action = 1).order_by('-id')
  return render(request,'admin/distributor_details.html',{'data':data})

def distributor_details_overview(request,id):
  data = Distributors_details.objects.get(id=id)
  return render(request,'admin/distributor_details_overview.html',{'data':data})

def dcompany_request(request):
  
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 0,reg_action='distributor').order_by('-id')
  return render(request,'distributor/dcompany_request.html',{'data':data,'distributor':distributor})

def dcompany_request_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_request_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_accept_company(request,id):
  data=company.objects.filter(id=id).update(Distributor_approval=1)
  
  return redirect('dcompany_request')
def distributor_reject_company(request,id):
  data=company.objects.get(id=id)
  data.user.delete()
  data.delete()
  return redirect('dcompany_request')

def dcompany_details(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  data = company.objects.filter(Distributors = distributor,Distributor_approval = 1,reg_action='distributor').order_by('-id')
  
  return render(request,'distributor/dcompany_details.html',{'data':data,'distributor':distributor})

def dcompany_details_overview(request,id):
  com = company.objects.get(id=id)
  allmodules=modules_list.objects.get(company=id)
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/dcompany_details_overview.html',{'company':com,'allmodules':allmodules,'distributor':distributor})

def distributor_profile(request):
  distributor =  Distributors_details.objects.get(user = request.user)
  return render(request,'distributor/distributor_profile.html',{'distributor':distributor})

# ========================================   ASHIKH V U (START) ======================================================

@login_required(login_url='login')
def item_create(request):
  item_units = UnitModel.objects.filter(user=request.user.id)
  return render(request,'company/item_create.html',{'item_units':item_units})

@login_required(login_url='login')
def items_list(request,pk):
  try:
    get_company_id_using_user_id = company.objects.get(user=request.user.id)
    all_items = ItemModel.objects.filter(company=get_company_id_using_user_id.id)
    if pk == 0:
      first_item = all_items.filter().first()
    else:
      first_item = all_items.get(id=pk)
    transactions = TransactionModel.objects.filter(user=request.user.id,item=first_item.id).order_by('-trans_created_date')
    check_var = 0
    if all_items == None or all_items == '' or first_item == None or first_item == '' or transactions == None or transactions == '':
      return render(request,'company/items_create_first_item.html')
    return render(request,'company/items_list.html',{'all_items':all_items,
                                                      'first_item':first_item,
                                                      'transactions':transactions,})
  except:
    return render(request,'company/items_create_first_item.html')

@login_required(login_url='login')
def item_create_new(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' or None :
      item_opening_stock = 0
      item_current_stock = 0
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '' or None:
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == ''  or None:
      item_min_stock_maintain = 0
    item_data = ItemModel(user=user,
                          company=company_user_data,
                          item_name=item_name,
                          item_hsn=item_hsn,
                          item_unit=item_unit,
                          item_taxable=item_taxable,
                          item_gst=item_gst,
                          item_igst=item_igst,
                          item_sale_price=item_sale_price,
                          item_purchase_price=item_purchase_price,
                          item_opening_stock=item_opening_stock,
                          item_current_stock=item_current_stock,
                          item_at_price=item_at_price,
                          item_date=item_date,
                          item_min_stock_maintain=item_min_stock_maintain)
    item_data.save()
    print(f'user : {user}\ncompany_user_data {company_user_data}')
    # print(f'item_name : {item_name}\nitem_hsn : {item_hsn}\nitem_unit : {item_unit}\nitem_taxable : {item_taxable}\n')
    # print(f'item_gst : {item_gst}\nitem_igst : {item_igst}\nitem_sale_price : {item_sale_price}\nitem_purchase_price : {item_purchase_price}\n')
    # print(f'item_opening_stock : {item_opening_stock}\nitem_at_price : {item_at_price}\nitem_date : {item_date}\nitem_min_stock_maintain : {item_min_stock_maintain}\n')
    print(f"----------\n\n\n")
    if request.POST.get('save_and_next'):
      return redirect('item_create')
    elif request.POST.get('save'):
      return redirect('items_list',pk=item_data.id)
  return redirect('item_create')


@login_required(login_url='login')
def item_delete(request,pk):
  get_company_id_using_user_id = company.objects.get(user=request.user.id)
  item_to_delete = ItemModel.objects.get(id=pk)
  item_to_delete.delete()
  return redirect('items_list',pk=0)


@login_required(login_url='login')
def item_view_or_edit(request,pk):
  item = ItemModel.objects.get(id=pk)
  item_units = UnitModel.objects.filter(user=request.user.id)
  return render(request,'company/item_view_or_edit.html',{'item':item,
                                                          'item_units':item_units,})

  
@login_required(login_url='login')
def item_unit_create(request):
  if request.method=='POST':
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_unit_name = request.POST.get('item_unit_name')
    unit_data = UnitModel(user=user,company=company_user_data,unit_name=item_unit_name)
    unit_data.save()
  return JsonResponse({'message':'asdasd'})

  
@login_required(login_url='login')
def item_update(request,pk):
  if request.method=='POST':
    item_data = ItemModel.objects.get(id=pk)
    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    item_name = request.POST.get('item_name')
    item_hsn = request.POST.get('item_hsn')
    item_unit = request.POST.get('item_unit')
    item_taxable = request.POST.get('item_taxable')
    item_gst = request.POST.get('item_gst')
    item_igst = request.POST.get('item_igst')
    if item_taxable == 'Non Taxable':
      item_gst = 'GST0[0%]'
      item_igst = 'IGST0[0%]'
    item_sale_price = request.POST.get('item_sale_price')
    item_purchase_price = request.POST.get('item_purchase_price')
    item_opening_stock = request.POST.get('item_opening_stock')
    item_current_stock = item_opening_stock
    if item_opening_stock == '' :
      item_opening_stock = 0
      item_current_stock = 0
    else:
      if int(item_opening_stock) > item_data.item_opening_stock:
        item_data.item_current_stock += (int(item_opening_stock) - item_data.item_opening_stock)
      else:
        item_data.item_current_stock -= (int(item_opening_stock) - item_data.item_opening_stock)
    item_at_price = request.POST.get('item_at_price')
    if item_at_price == '':
      item_at_price =0
    item_date = request.POST.get('item_date')
    item_min_stock_maintain = request.POST.get('item_min_stock_maintain')
    if item_min_stock_maintain == '':
      item_min_stock_maintain = 0

    item_data.user = user
    item_data.company_user_data = company_user_data
    item_data.item_name = item_name
    item_data.item_hsn = item_hsn
    item_data.item_unit = item_unit
    item_data.item_taxable = item_taxable
    item_data.item_gst = item_gst
    item_data.item_igst = item_igst
    item_data.item_sale_price = item_sale_price
    item_data.item_purchase_price = item_purchase_price
    item_data.item_opening_stock = item_opening_stock
    item_data.item_current_stock = int(item_current_stock)
    item_data.item_at_price = item_at_price
    item_data.item_date = item_date
    item_data.item_min_stock_maintain = item_min_stock_maintain

    item_data.save()
    print('\nupdated')
  # return redirect('item_view_or_edit',pk)
  return redirect('items_list',pk=item_data.id)

  
@login_required(login_url='login')
def item_search_filter(request):
  search_string = request.POST.get('searching_item')
  items_filtered = ItemModel.objects.filter(user=request.user.id)
  items_filtered = items_filtered.filter(Q(item_name__icontains=search_string))
  item_unit_name = request.POST.get('item_unit_name')
  return TemplateResponse(request,'company/item_search_filter.html',{'all_items':items_filtered})


@login_required(login_url='login')
def item_get_detail(request,pk):
  item = ItemModel.objects.get(id=pk)
  transactions = TransactionModel.objects.filter(user=request.user.id,item=item.id).order_by('-trans_created_date')
  return TemplateResponse(request,'company/item_get_detail.html',{"item":item,
                                                                  'transactions':transactions,})

  
@login_required(login_url='login')
def item_get_details_for_modal_target(request,pk):
  item = ItemModel.objects.get(id=pk)
  return TemplateResponse(request,'company/item_get_details_for_modal_target.html',{"item":item,})


@login_required(login_url='login')
def ajust_quantity(request,pk):
  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    trans_adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    print(f'the quantity : {trans_current_qty}')
    item.item_current_stock = trans_adjusted_qty
    item.save()
    transaction_data = TransactionModel(user=user,
                                        company=company_user_data,
                                        item=item,
                                        trans_type=trans_type,
                                        trans_user_name=trans_user_name,
                                        trans_date=trans_date,
                                        trans_qty=trans_qty,
                                        trans_current_qty=trans_current_qty,
                                        trans_adjusted_qty=trans_adjusted_qty,)
    transaction_data.save()
  return redirect('items_list',pk=item.id)


@login_required(login_url='login')
def transaction_delete(request,pk):
  transaction = TransactionModel.objects.get(id=pk)
  item = ItemModel.objects.get(id=transaction.item_id)
  print(transaction.trans_type)
  if transaction.trans_type=='add stock':
    print('add')
    item.item_current_stock = item.item_current_stock - transaction.trans_qty
    print(item.item_name)
    print(item.item_current_stock)
    print(item.item_current_stock)
    print(transaction.trans_qty)
    print(item.item_current_stock - transaction.trans_qty)
  else:
    print('reduce')
    item.item_current_stock = item.item_current_stock + transaction.trans_qty
  item.save()
  transaction.delete()
  return redirect('items_list',pk=item.id)

  
@login_required(login_url='login')
def item_transaction_view_or_edit(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  print('enterd')
  return TemplateResponse(request,'company/item_transaction_view_or_edit.html',{"item":item,
                                                                                "transaction":transaction,})


@login_required(login_url='login')
def update_adjusted_transaction(request,pk,tran):
  item = ItemModel.objects.get(id=pk)
  transaction = TransactionModel.objects.get(id=tran)
  if request.method=='POST':
    item = ItemModel.objects.get(id=pk)

    user = User.objects.get(id=request.user.id)
    company_user_data = company.objects.get(user=request.user.id)
    trans_type_check_checked = request.POST.get('trans_type')
    if trans_type_check_checked == 'on':
      trans_type = 'reduce stock'
      trans_qty = request.POST.get('reduced_qty')
    else:
      trans_type = 'add stock'
      trans_qty = request.POST.get('added_qty')
    trans_user_name = user.first_name
    trans_date = request.POST.get('trans_date')

    adjusted_qty= request.POST.get('adjusted_qty')
    trans_current_qty = request.POST.get('item_qty')
    if transaction.trans_type == 'reduce stock':
      if trans_type == 'reduce stock':
        print('reduce to reduce')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  - transaction.trans_qty)
      else:
        print('reduce to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock + transaction.trans_qty + int(trans_qty)
    else:
      if trans_type == 'reduce stock':
        print('add to red')
        item.item_current_stock = item.item_current_stock - (int(trans_qty)  + transaction.trans_qty)
      else:
        print('add to add')
        print(f'{trans_qty}-{transaction.trans_qty}={((int(trans_qty)  - transaction.trans_qty))}')
        item.item_current_stock = item.item_current_stock - transaction.trans_qty + int(trans_qty)
    # item.item_opening_stock = adjusted_qty
    item.save()
    transaction.trans_type =trans_type
    transaction.trans_date=trans_date
    transaction.trans_qty =trans_qty
    transaction.trans_current_qty=trans_current_qty
    transaction.save()
  return redirect('items_list',pk=item.id)
  
@login_required(login_url='login')
def item_delete_open_stk(request,pk):
  item = ItemModel.objects.get(id=pk)
  if item.item_opening_stock > item.item_current_stock:
    item.item_current_stock =item.item_opening_stock - item.item_current_stock
  else:
    item.item_current_stock =item.item_current_stock - item.item_opening_stock
  # item.item_current_stock =  item.item_opening_stock - item.item_current_stock
  item.item_opening_stock = 0
  # print(f'{item.item_current_stock }={item.item_opening_stock}-{item.item_current_stock}')
  item.save()
  return redirect('items_list',pk=item.id)
  
# ========================================   ASHIKH V U (END) ======================================================

#_________________Parties(new)_______________Antony Tom_________


def add_parties(request):
  return render(request, 'company/add_parties.html')


def save_parties(request):
    if request.method == 'POST':
        Company = company.objects.get(user=request.user)
        user_id = request.user.id
        
        party_name = request.POST['partyname']
        gst_no = request.POST['gstno']
        contact = request.POST['contact']
        gst_type = request.POST['gst']
        state = request.POST['state']
        address = request.POST['address']
        email = request.POST['email']
        openingbalance = request.POST.get('balance', '')
        payment = request.POST.get('paymentType', '')
        creditlimit = request.POST.get('creditlimit', '')
        current_date = request.POST['currentdate']
        End_date = request.POST.get('enddate', None)
        additionalfield1 = request.POST['additionalfield1']
        additionalfield2 = request.POST['additionalfield2']
        additionalfield3 = request.POST['additionalfield3']
        user=User.objects.get(id=user_id)
        comp=Company
        if (
          not party_name
          
      ):
          return render(request, 'add_parties.html')

        part = party(party_name=party_name, gst_no=gst_no,contact=contact,gst_type=gst_type, state=state,address=address, email=email, openingbalance=openingbalance,payment=payment,
                       creditlimit=creditlimit,current_date=current_date,End_date=End_date,additionalfield1=additionalfield1,additionalfield2=additionalfield2,additionalfield3=additionalfield3,user=user,company=comp)
        part.save() 

        if 'save_and_new' in request.POST:
            
            return render(request, 'company/add_parties.html')
        else:
          
            return redirect('view_parties')

    return render(request, 'company/add_parties.html')  


def view_parties(request):
  Company = company.objects.get(user=request.user.id)
  user_id = request.user.id
  Party=party.objects.filter(company=Company.id)
  return render(request, 'company/view_parties.html',{'Company':Company,'user_id':user_id,'Party':Party})


def view_party(request,id):
  Company = company.objects.get(user=request.user)
  user_id = request.user.id
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(company=Company.id)
  return render(request, 'company/view_party.html',{'Company':Company,'user_id':user_id,'Party':Party,'getparty':getparty})

def edit_party(request,id):
  Company = company.objects.get(user=request.user)
  user_id = request.user.id
  getparty=party.objects.get(id=id)
  Party=party.objects.filter(user=request.user)
  return render(request, 'company/edit_party.html',{'Company':Company,'user_id':user_id,'Party':Party,'getparty':getparty})


def edit_saveparty(request, id):
    Party=party.objects.filter(user=request.user)
    user_id = request.user.id
    getparty = party.objects.get(id=id)
    Company = company.objects.get(user=request.user)

    if request.method == 'POST':
        getparty.party_name = request.POST.get('partyname')
        getparty.gst_no = request.POST.get('gstno')
        getparty.contact = request.POST['contact']
        getparty.gst_type = request.POST['gst']
        getparty.state = request.POST['state']
        getparty.address = request.POST['address']
        getparty.email = request.POST['email']
        getparty.openingbalance = request.POST['balance']
        getparty.payment = request.POST.get('paymentType')
        getparty.creditlimit = request.POST['creditlimit']
        getparty.current_date = request.POST['currentdate']
        getparty.additionalfield1 = request.POST['additionalfield1']
        getparty.additionalfield2 = request.POST['additionalfield2']
        getparty.additionalfield3 = request.POST['additionalfield3']

        getparty.save()

        return redirect('view_party', id=getparty.id)

    return render(request, 'edit_party.html', {'getparty': getparty, 'Party': Party, 'Company': Company,'user_id':user_id})


def deleteparty(request,id):
    Party=party.objects.get(id=id)
    Party.delete()
    return redirect('view_parties')

#End

@login_required(login_url='login')
def adminhome(request):
 
  
  
  return render(request, 'admin/adminhome.html')


#******************************************   ASHIKH V U (start) ****************************************************

from django.http import HttpResponse
import re

# account number validation
def validate_bank_account_number(acc_num):
  regex='^[0-9]{9,18}'
  if re.match(regex,acc_num):
    return True
  else:
    return False

# ifsc code validaion
def validate_ifsc(ifsc_code):
    regex = re.compile(r'^[A-Za-z]{4}\d{7}$')
    if regex.match(ifsc_code):
        return True
    else:
        return False

#@login_required(login_url='login')
def account_num_check(request):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def account_num_check_for_edit(request,pk):
  if request.method=='POST':
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    account_num_valid = validate_bank_account_number(account_num)
    if account_num_valid:
      if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
        return HttpResponse('<small><span class="tr fs-2">Account Number already excist</span></small>')
      else:
        return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">Account Number is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_ifsc_check (request):
  if request.method=='POST':
    bank_ifsc = request.POST.get('ifsc')
    print(bank_ifsc)
    ifsc_valid = validate_ifsc(bank_ifsc)
    if ifsc_valid:
      return HttpResponse('')
    else:
      return HttpResponse('<small><span class="tr fs-2">IFSC Code is not valid</span></small>')
  return HttpResponse('')

#@login_required(login_url='login')
def bank_create(request):
  print('asdasd')
  try:
    staff_id = request.session['staff_id']
    print(staff_id)
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})
  except:
    user = User.objects.get(id=request.user.id)
    get_company_id_using_user_id = company.objects.get(user=user)
    # permission
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    # permission
    return render(request,'company/bank_create.html',{"allmodules":allmodules})


#@login_required(login_url='login')
def banks_list(request,pk):
  
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
 

  try:
    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
    if pk == 0:
      first_bank = all_banks.first()
      print(all_banks)
      return redirect('banks_list',pk=first_bank.id)
    else:
      bank = all_banks.get(id=pk)
      transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
      transactions = transactions_all.filter(Q(from_here=pk) | Q(to_here=pk))
      tr_history = BankTransactionHistory.objects.filter().order_by('date')
    if all_banks.exists():
      open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
      
      if bank.open_balance:
        total = bank.open_balance
      else:
        total = 0
      for i in transactions:
        if i.type == "Cash Withdraw":
          total = total - i.amount
        elif  i.type == 'Adjustment Reduce':
          total = total - i.amount
        elif i.from_here == bank:
          total = total - i.amount
        else:
          total = total + i.amount
        i.current_amount = total
      
      return render(request,'company/banks_list.html',{"allmodules":allmodules,
                                                      "all_banks":all_banks,
                                                      "bank":bank,
                                                      "transactions":transactions,
                                                      "tr_history":tr_history,
                                                      "open_bal_last_edited":open_bal_last_edited,
                                                      "staff":staff}) 
    else:
      return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
  except:
    return render(request,'company/bank_create_first_bank.html',{"allmodules":allmodules,'staff':staff}) 
    

#@login_required(login_url='login')
def get_bank_to_bank(request):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_bank_to_cash(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_bank_to_cash.html',{'banks':banks})

#@login_required(login_url='login')
def get_cash_to_bank(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_cash_to_bank.html',{'banks':banks})

#@login_required(login_url='login')
def get_adjust_bank_balance(request):
  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  return TemplateResponse(request,'company/bank_adjust_bank_balance.html',{'banks':banks})

#@login_required(login_url='login')
def bank_create_new(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    print(get_company_id_using_user_id)
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(company=get_company_id_using_user_id.id).filter(bank_name=bank_name,user=user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
      
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data = BankModel(user=user,
                                company=get_company_id_using_user_id,
                                bank_name=bank_name,
                                account_num=account_num,
                                ifsc=ifsc,
                                branch_name=branch_name,
                                upi_id=upi_id,
                                as_of_date=as_of_date,
                                card_type=card_type,
                                open_balance=open_balance,
                                current_balance=open_balance,
                                created_by=user.first_name)
          bank_data.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK CREATION : "+bank_data.bank_name.upper(),
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                              bank=bank_data,
                                              action="BANK OPEN BALANCE CREATED",
                                              done_by_name=staff.first_name,
                                              done_by=staff)
          tr_history.save()
          if request.POST.get('save_and_next'):
            messages.success(request,'Bank created successfully')
            return redirect('bank_create')
          else:
            return redirect('banks_list',pk=bank_data.id)
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)

#@login_required(login_url='login')
def bank_delete(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.delete()
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def bank_view_or_edit(request,pk):
  bank = BankModel.objects.get(id=pk)
  return render(request,'company/bank_view_or_edit.html',{"bank":bank})

#@login_required(login_url='login')
def bank_update(request,pk):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
    
    bank_data = BankModel.objects.get(id=pk)

    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    bank_name = request.POST.get('bank_name')
    account_num = request.POST['account_num']
    if BankModel.objects.exclude(id=pk).filter(bank_name=bank_name,user=request.user.id,account_num=account_num).exists():
      parmission_var = 0
    else:
      parmission_var = 1
    if validate_bank_account_number(account_num):
      parmission_var1 = 1
    else:
      parmission_var1 = 0
    ifsc = request.POST.get('ifsc')
    if validate_ifsc(ifsc):
      parmission_var2 = 1
    else:
      parmission_var2 = 0
    branch_name = request.POST['branch_name']
    upi_id = request.POST.get('upi_id')
    as_of_date = request.POST['as_of_date']
    card_type = request.POST.get('card_type')
    open_balance = request.POST['open_balance']
    
    if open_balance == '' or open_balance == None:
      open_balance = 0
    if card_type == "CREDIT":
      open_balance = int(open_balance)*-1
    if parmission_var == 1:
      if parmission_var1 == 1:
        if parmission_var2 == 1:
          bank_data.user = user
          bank_data.company = get_company_id_using_user_id
          bank_data.bank_name = bank_name
          bank_data.account_num = account_num
          bank_data.ifsc = ifsc
          bank_data.branch_name = branch_name
          bank_data.upi_id = upi_id
          bank_data.as_of_date = as_of_date
          bank_data.card_type = card_type

          if int(bank_data.open_balance) < int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance) + (int(open_balance) - int(bank_data.open_balance))
          elif int(bank_data.open_balance) == int(open_balance):
            bank_data.current_balance = int(open_balance)
          elif int(bank_data.open_balance) > int(open_balance):
            bank_data.current_balance = int(bank_data.current_balance)- (int(bank_data.open_balance) - int(open_balance))

          if bank_data.open_balance != open_balance:
            validity = True
          else:
            validity = False
          old_val = bank_data.open_balance

          bank_data.open_balance = open_balance
          bank_data.user = user
          bank_data.save()

          if validity == True:
            tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank=bank_data,
                                          action="BANK OPEN BALANCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
            tr_history.save()
        else:
          messages.error(request,'IFSC CODE is not valid')
          return redirect('bank_create')
      else:
        messages.error(request,'Account number is not valid')
        return redirect('bank_create')
    else:
      messages.error(request,'Account number already exist')
      return redirect('bank_create')
  return redirect('banks_list',pk=bank_data.id)


#@login_required(login_url='login')
def bank_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "BANK TO BANK"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date')
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        to_here=to_here,
                                        type=type,
                                        date=date,
                                        name=name,
                                        amount=amount,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO BANK TRANSACTION CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def bank_to_cash_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    type = "Cash Withdraw"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    bank1.current_balance -= int(amount)
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def cash_to_bank_transaction_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select2 = request.POST.get('to_here')
    to_here = BankModel.objects.get(id=select2)
    type = "Cash Deposit"
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank2 = BankModel.objects.get(id=to_here.id)
    bank2.current_balance += int(amount)
    bank2.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        to_here=to_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank2,
                                        bank_trans=transaction_data,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=to_here.id)


#@login_required(login_url='login')
def get_adjust_bank_balance_create(request):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    select1 = request.POST.get('from_here')
    from_here = BankModel.objects.get(id=select1)
    inc_red = request.POST.get('inc_red')
    name = staff.first_name
    amount = request.POST.get('amount')
    date = request.POST.get('date') 
    
    bank1 = BankModel.objects.get(id=from_here.id)
    if inc_red == 'Increase Balance':
      bank1.current_balance += int(amount) 
      type = "Adjustment Increase"
    else:
      bank1.current_balance -= int(amount)
      type = "Adjustment Reduce"
    bank1.save()

    transaction_data = BankTransactionModel(user = user,
                                        company=get_company_id_using_user_id,
                                        from_here=from_here,
                                        type=type,
                                        name=name,
                                        amount=amount,
                                        date=date,
                                        last_action='CREATED',
                                        by = staff.first_name,
                                        )
    transaction_data.save()
    tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=bank1,
                                        bank_trans=transaction_data,
                                        action="BANK BALANCE "+type.upper()+" CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
    tr_history.save()
  return redirect('banks_list',pk=from_here.id)

#@login_required(login_url='login')
def delete_bank_open_balance(request,pk):
  bank = BankModel.objects.get(id=pk)
  bank.current_balance = bank.current_balance - bank.open_balance
  bank.open_balance = 0
  bank.save()
  if 'banks_list' in request.META.get('HTTP_REFERER',None):
    return redirect('banks_list',pk=pk)
  else:
    return redirect('bank_transaction_statement',bank_id=pk)

#@login_required(login_url='login')
def delete_bank_transaction(request,pk,bank_id):
  print(pk,bank_id)
  try:
    pk = request.POST.get('pk')
    bank_id = request.POST.get('bank_id')
    print(pk,bank_id)
  except:
    pk=pk
    bank_id=bank_id

  try:
    trans = BankTransactionModel.objects.get(id=pk)
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('enterd')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Cash Deposit' or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      bank2.current_balance -= trans.amount
      bank2.save()
      trans.delete()
      print('entered')
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance -= trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      bank1.current_balance += trans.amount
      bank1.save()
      trans.delete()
      return redirect('banks_list',pk=bank_id)
  except:
    return redirect('banks_list',pk=bank_id)
  return redirect('banks_list',pk=0)

#@login_required(login_url='login')
def view_or_edit_bank_transaction(request,pk,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

  transaction = BankTransactionModel.objects.get(id=pk)
  banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)
  bank = BankModel.objects.get(id=bank_id)
  if transaction.type == "BANK TO BANK" or transaction.type == 'Bank to bank':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Withdraw' or transaction.type == 'Cash withdraw' or transaction.type == 'CASH WITHDRAW':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_to_cash_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Cash Deposit' or transaction.type == 'Cash deposit' or transaction.type == 'CASH DEPOSIT':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/cash_to_bank_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})
  elif transaction.type == 'Adjustment Increase' or transaction.type == 'Adjustment increase' or transaction.type == 'Adjustment Reduce' or transaction.type == 'Adjustment reduce' or transaction.type == 'ADJUSTMENT INCREASE' or transaction.type == 'ADJUSTMENT REDUCE':
    return TemplateResponse(request,'company/bank_transaction_view_or_edit/bank_adjust_bank_balance_view_or_edit.html',{"transaction":transaction,"banks":banks,"bank":bank})

#@login_required(login_url='login')
def update_bank_transaction(request,pk,bank_id):
  if request.method=="POST":

    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    amount = request.POST.get('amount')
    date = request.POST.get('date')
    trans = BankTransactionModel.objects.get(id=pk)
    trans.date = date
    if trans.type == 'BANK TO BANK':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      old_amount = trans.amount
      if old_amount != amount:
        validity =True
      else:
        validity =False
      trans.amount = amount
      trans.save()
      if validity == True:
        tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
        tr_history.save()
        trans.last_action='UPDATED'
        trans.by = staff.first_name
        trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=bank_id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Withdraw' or trans.type == 'CASH WITHDRAW':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK TO CASH TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Cash Deposit'  or trans.type == 'CASH DEPOSIT':
      bank2 = BankModel.objects.get(id=trans.to_here.id)
      if trans.amount > int(amount):
        bank2.current_balance -= (trans.amount-int(amount))
      else:
        bank2.current_balance += (int(amount)-trans.amount)
      bank2.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="CASH TO BANK TRANSACTION UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.to_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Increase' or trans.type == 'ADJUSTMENT INCREASE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance -= (trans.amount-int(amount))
      else:
        bank1.current_balance += (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT INCREASE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    elif trans.type == 'Adjustment Reduce' or trans.type == 'ADJUSTMENT REDUCE':
      bank1 = BankModel.objects.get(id=trans.from_here.id)
      if trans.amount > int(amount):
        bank1.current_balance += (trans.amount-int(amount))
      else:
        bank1.current_balance -= (int(amount)-trans.amount)
      bank1.save()
      trans.amount = amount
      trans.save()
      tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                          bank_trans=trans,
                                          action="BANK BALANCE ADJUSTMENT REDUCE UPDATED",
                                          done_by_name=staff.first_name,
                                          done_by=staff)
      tr_history.save()
      trans.last_action='UPDATED'
      trans.by = staff.first_name
      trans.save()
      if 'banks_list' in request.META.get('HTTP_REFERER',None):
        return redirect('banks_list',pk=trans.from_here.id)
      else:
        return redirect('bank_transaction_statement',bank_id=bank_id)
    return redirect('banks_list',pk=0)
  return redirect('banks_list',pk=0)

from openpyxl import load_workbook
from django.utils import timezone

#@login_required(login_url='login')
def import_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('banks_list',pk=pk)

#@login_required(login_url='login')
def import_statement_from_excel(request,pk):
    current_datetime = timezone.now()
    date =  current_datetime.date()

    try:
      if request.method == "POST" and 'excel_file' in request.FILES:
        
        staff_id = request.session['staff_id']
        staff =  staff_details.objects.get(id=staff_id)
        get_company_id_using_user_id = company.objects.get(id=staff.company.id)
        user = get_company_id_using_user_id.user
        allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

        excel_file = request.FILES['excel_file']

        wb = load_workbook(excel_file)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE = row
            # TYPE, FROM, TO,NAME,DATE,AMOUNT,BALANCE,ACTION,BY = row

            if TYPE != None:
              TYPE = TYPE.upper()
            
            if AMOUNT != None:
              AMOUNT = AMOUNT.replace(' ','')
              AMOUNT = AMOUNT.replace('₹','')
              AMOUNT = AMOUNT.replace('-','')
              AMOUNT = AMOUNT.replace('+','')
              AMOUNT = int(float(AMOUNT))

            print(f'{TYPE}  {FROM}  {TO}    {NAME}  {DATE}  {AMOUNT}')
            
            if TYPE == "BANK TO BANK" or TYPE == 'Bank to bank':
              from_here = BankModel.objects.get(id=int(FROM))
              to_here = BankModel.objects.get(id=int(TO))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  to_here=to_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              to_here.current_balance += AMOUNT
              to_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  bank_trans=transaction,
                                                  action="BANK TO BANK TRANSACTION CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
              
            elif TYPE == 'Open. Balance' or TYPE == 'OPEN. BALANCE':
              from_here = BankModel.objects.get(id=int(FROM))
              if from_here.open_balance > AMOUNT:
                from_here.current_balance += from_here.open_balance - AMOUNT
              else:
                from_here.current_balance -= from_here.open_balance - AMOUNT
              from_here.open_balance = AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                                  bank=from_here,
                                                  action="BANK OPEN BALANCE CREATED",
                                                  done_by_name=staff.first_name,
                                                  done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Withdraw' or TYPE == 'Cash withdraw' or TYPE == 'CASH WITHDRAW':
              from_here = BankModel.objects.get(id=int(FROM))
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              from_here.current_balance -= AMOUNT
              from_here.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK TO CASH TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Cash Deposit' or TYPE == 'Cash deposit' or TYPE == 'CASH DEPOSIT':
              to_here = BankModel.objects.get(id=int(TO))
              to_here.current_balance += AMOUNT
              to_here.save()

              transaction = BankTransactionModel(user = user,
                                                  company=get_company_id_using_user_id,
                                                  to_here=to_here,
                                                  type=TYPE,
                                                  amount=AMOUNT,
                                                  date=DATE,
                                                  last_action='CREATED',
                                                  by = staff.first_name,
                                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=to_here,
                                        bank_trans=transaction,
                                        action="CASH TO BANK TRANSACTION CREATED",
                                        date=date,
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Increase' or TYPE == 'Adjustment increase' or TYPE == 'ADJUSTMENT INCREASE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance += AMOUNT
              from_here.save()
              transaction =BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
            elif TYPE == 'Adjustment Reduce' or TYPE == 'Adjustment reduce' or TYPE == 'ADJUSTMENT REDUCE':
              from_here = BankModel.objects.get(id=int(FROM))
              from_here.current_balance -= AMOUNT
              from_here.save()
              transaction = BankTransactionModel(user = user,
                                  company=get_company_id_using_user_id,
                                  from_here=from_here,
                                  type=TYPE,
                                  amount=AMOUNT,
                                  date=DATE,
                                  last_action='CREATED',
                                  by = staff.first_name,
                                  )
              transaction.save()
              tr_history = BankTransactionHistory(company=get_company_id_using_user_id,
                                        bank=from_here,
                                        bank_trans=transaction,
                                        action="BANK BALANCE ADJUSTMENT REDUCE CREATED",
                                        done_by_name=staff.first_name,
                                        done_by=staff)
              tr_history.save()
    except:
      messages.warning(request,"Table field is missing / you are importing the wrong File.")
    return redirect('bank_transaction_statement',bank_id=pk) 

#@login_required(login_url='login')
def transaction_history(request,pk,bank_id):
    
    staff_id = request.session['staff_id']
    staff =  staff_details.objects.get(id=staff_id)
    get_company_id_using_user_id = company.objects.get(id=staff.company.id)
    user = get_company_id_using_user_id.user
    allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')

    all_banks = BankModel.objects.filter(company=get_company_id_using_user_id.id)

    # tr_history1 = BankTransactionHistory.objects.filter(action__contains='BANK CREATION',bank=bank_id)
    tr_history2 = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED'),bank=bank_id)    
    tr_history = BankTransactionHistory.objects.filter(bank_trans=pk)
    if pk != 0:
      # tr_historys = tr_history | tr_history1
      tr_historys = tr_history
    else:
      # tr_historys = tr_history1 | tr_history1 | tr_history2
      tr_historys =  tr_history | tr_history2
    # print(tr_history)
    
    return render(request,'company/bank_transaction_history.html',{"allmodules":allmodules,
                                                                   "all_banks":all_banks,
                                                                    "tr_historys":tr_historys,
                                                                    "bank_id":bank_id,
                                                                    "staff":staff})


#@login_required(login_url='login')
def bank_transaction_statement(request,bank_id):

  staff_id = request.session['staff_id']
  staff =  staff_details.objects.get(id=staff_id)
  get_company_id_using_user_id = company.objects.get(id=staff.company.id)
  user = get_company_id_using_user_id.user
  allmodules= modules_list.objects.get(company=get_company_id_using_user_id,status='New')
  bank = BankModel.objects.get(id=bank_id)

  transactions_all = BankTransactionModel.objects.filter(company=get_company_id_using_user_id.id)
  transactions = transactions_all.filter(Q(from_here=bank_id) | Q(to_here=bank_id))
  tr_history = BankTransactionHistory.objects.filter().order_by('date')

  open_bal_last_edited = BankTransactionHistory.objects.filter(Q(action__contains='BANK OPEN BALANCE CREATED') | Q(action__contains='BANK OPEN BALANCE UPDATED')).last()
  
  if bank.open_balance:
    total = bank.open_balance
  else:
    total = 0
  for i in transactions:
    if i.type == "Cash Withdraw":
      total = total - i.amount
    elif  i.type == 'Adjustment Reduce':
      total = total - i.amount
    elif i.from_here == bank:
      total = total - i.amount
    else:
      total = total + i.amount
    i.current_amount = total

  return render(request,'company/bank_transaction_statement.html',{"allmodules":allmodules,
                                                  "bank":bank,
                                                  "transactions":transactions,
                                                  "tr_history":tr_history,
                                                  "open_bal_last_edited":open_bal_last_edited,
                                                  "staff":staff})

#******************************************   ASHIKH V U (end) ****************************************************
#********************************************keerthana***********************************************************

@login_required(login_url='login')
def sales_first(request):
  return render(request,'company/salesfirst.html')

@login_required(login_url='login')
def create_sale(request):
  return render(request,'company/create_sale.html')
